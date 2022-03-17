#!/usr/bin/python
# coding:utf-8
"""
Create Time: 2022/3/14 16:36
Author: yinyilin
#!/usr/bin/env python
"""

# 导入工具类
from openpyxl import load_workbook
# 定义打开Excel类
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:

    """操作Excel"""
    def __init__(self, file):
        """初始化，定义路径"""
        self.file = file

    def open_sheet(self, sheet_name) -> Worksheet:
        """
        打开表单
        在函数或者方法的后面加 -> 类型：表示此函数返回值是这样的类型
        """
        wb = load_workbook(self.file) # 打开Excel
        # 打开表单
        sheet = wb[sheet_name]
        wb.close()
        return sheet

    def header(self,sheet_name):
        """获取数据"""
        # 打开表单
        sheet = self.open_sheet(sheet_name)
        header = []
        for i in sheet[1]:
            header.append(i.value)  #获取值
        return header

    def read(self, sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)

        rows = list(sheet.rows)
        # rows = list(sheet.rows)[1:] # 得到一个生成器，用list转换

        data = [] # 定义列表，存储数据列表
        for row in rows[1:]:
            row_data = [] # 获取数据列表
            for cell in row: # 加载
                # print(cell.value) # 获取数据
                row_data.append(cell.value) # 添加数据进去
                # 列表转换成字典:要和header去zip
                data_dict = dict(zip(self.header(sheet_name),row_data))
            data.append(data_dict)
        return data


    # def write(self,shett_name,row,column,data):
    #     """写入Excel数据"""
    #     sheet = self.open_sheet(shett_name)
    @staticmethod
    def write(file,sheet_name,row,column,data):
        """写入Excel数据"""
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row,column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()